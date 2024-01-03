import sys
import time

import openpyxl

from helper.Non_selenium_hilfs_klassen.ErstelleZeitstempelAlsString import ErstelleZeitstempelAlsString
from helper.Non_selenium_hilfs_klassen.LeseAbsPfadVonDateienUndOrdnern import LeseAbsPfadVonDateienUndOrdnern
from helper.Non_selenium_hilfs_klassen.SchreibeTestCaseUndSchirttNameInFreieZeile import \
    SchreibeTestCaseUndSchirttNameInFreieZeile
from helper.Selenium_hilfs_klassen.ErstelleUndSpeicherScreenshot import ErstelleUndSpeicherScreenshot


class ReporteTestCaseUndStepErgebnisseInXlsx:

    """
    ####################################################################################################################
    Autor: [Rene Erdmann]
    Erstellt am: [20.09.2023]
    ####################################################################################################################
    """

    def __init__(self,
                 driver,
                 release_name,
                 test_name,
                 test_ausfuerungs_tabelle,
                 test_ausfuerungs_tabelle_tab_name,
                 abs_pfad_zum_test_ausfuerhungs_ordner
                 ):

        self.driver = driver
        self.test_ausfuerungs_tabelle = test_ausfuerungs_tabelle
        self.release_name = release_name
        self.ergebnis_xlsx_tab_sheet_name = test_ausfuerungs_tabelle_tab_name
        self.abs_pfad_zum_test_ausfuerhungs_ordner = abs_pfad_zum_test_ausfuerhungs_ordner

        current_path_obj = LeseAbsPfadVonDateienUndOrdnern()
        self.pfad_zur_test_ausfuehrungs_tabelle = current_path_obj.get_abs_path_to_bund_id_test_lab_execution_xlsx(
            test_execution_xlsx_name=test_ausfuerungs_tabelle
        )

        self.start_splalte_der_testausfuerhung = 4

        self.workbook = openpyxl.load_workbook(self.pfad_zur_test_ausfuehrungs_tabelle)

        self.ergebnis_array = []
        sheet_name = 'Ausführung - Auto Test BundID'  # Passe den Namen des Arbeitsblatts an
        self.sheet = self.workbook[sheet_name]



        self.zeitstempel_doppelpunkt = ErstelleZeitstempelAlsString. \
            aktueller_zeitstempel_jahr_monat_tag_stunde_minute_tz_doppelpunkt()

        self.spalten_namen = []
        self.methoden_name = test_name

        try:
            self.sheet = self.workbook[self.ergebnis_xlsx_tab_sheet_name]
        except:
            print(
                " ---------- ERROR: Sheet konnte nicht geladen werden - sheet name muss lauten [Ausführung"
                " - Auto Test BundID] ")

        # Iteriere durch die erste Spalte der Ergebnis xlsx bis eine leere Spalte gefunden wird
        for cell in self.sheet[1]:
            if cell.value is None:
                self.erste_leere_zelle_der_zeile_eins = cell.column
                break

        self.test_run_nummer = self.erste_leere_zelle_der_zeile_eins

        if self.erste_leere_zelle_der_zeile_eins > self.start_splalte_der_testausfuerhung:

            # Diese Bedingung trifft zu, schon mehr als 1 Test-Run gemacht wurde.
            # Wert aus der Zelle auslesen
            name_der_letzten_execution = self.sheet.cell(row=1, column=self.erste_leere_zelle_der_zeile_eins - 1).value

            # Extrahiere die Zahl aus dem String "Release_7_0 : run-1" (Test-Run Bezeichnung)
            run_number = name_der_letzten_execution.split(" ")[-1].split("-")[-1]

            # Estelle Bezeichnung für neuen Test-Lauf
            neuer_test_run = self.release_name + " : run-" + str(int(run_number) + 1)

            # Nummer des Test-Run in zeile 1 eintragen
            self.sheet.cell(row=1, column=self.erste_leere_zelle_der_zeile_eins, value=neuer_test_run)

            # Gebe den bezeichner des Test-Runs aus
            print("Starten des Test-Runs...                     :", neuer_test_run)

        elif self.erste_leere_zelle_der_zeile_eins == self.start_splalte_der_testausfuerhung:

            # Diese Bedingung trifft zu, wenn noch kein Test-Run gemacht wurde
            # [5] da sie Test-Case Ausführung erst ab Spalte 5 beginnt (E).
            # Wenn kein Test-Run gemacht wurde, wird initial der run-1 gesetzt

            self.sheet.cell(row=1, column=self.erste_leere_zelle_der_zeile_eins, value=self.release_name + " : run-1")

        for i in range(26):
            self.spalten_namen.append(chr(65 + i))  # Buchstaben von A bis Z (ASCII-Werte von 65 bis 90)

        for i in range(26):
            for j in range(26):
                self.spalten_namen.append(chr(65 + i) + chr(65 + j))  # Kombinationen von AA bis GZ

        self.test_run_number = str((self.test_run_nummer - self.start_splalte_der_testausfuerhung) + 1)

    def get_test_run_nummer(self):

        return str((self.test_run_nummer - self.start_splalte_der_testausfuerhung) + 1)

    def trage_step_ergebnis_ein(self,
                                schritt_oder_gesamter_testfall,
                                erstelle_screenshot=True,
                                ergebnis=None,
                                ):


        self.test_case_war_existent = False

        if type(schritt_oder_gesamter_testfall) is int:

            schritt_oder_gesamter_testfall = "Step " + str(schritt_oder_gesamter_testfall)

        # Eingabe eines Direkten-Ergebnisses für Den Test-Schritt ######################################################
        # Wenn man zum Beispiel schon das aktuelle Ergebnis kennt un des einfach direkt in der excel speichern will.
        if ergebnis:

            if ergebnis.upper() == "PASSED":

                step_result = f"PASSED({self.zeitstempel_doppelpunkt})"
                self.ergebnis_array.append("PASSED")

            if ergebnis.upper() == "FAILED":

                step_result = f"FAILED({self.zeitstempel_doppelpunkt})"
                self.ergebnis_array.append("FAILED")

                if erstelle_screenshot:

                    erstelle_und_speichere_screen_shot = ErstelleUndSpeicherScreenshot(
                        driver=self.driver,
                        abs_pfad_zum_test_ausfuerhungs_ordner=self.abs_pfad_zum_test_ausfuerhungs_ordner,
                        test_run_number=self.test_run_number
                    )
                    erstelle_und_speichere_screen_shot.einzelner_screenshot_der_webseite(
                        step_number=schritt_oder_gesamter_testfall,
                        screen_shot_name="FAILED"
                    )
                    time.sleep(2)

        count_rows = 1
        count_empty_rows = 1

        for row in self.sheet.iter_rows(values_only=True):

            # überprüft in welcher Zeile der Spalte A sich der neue Testfall-Name un der TestSchritt befinden.
            # In diese Zeile wird dann das neuset Ergebnis eingetragen
            if row[0] == self.methoden_name and row[1] == schritt_oder_gesamter_testfall:

                # Wert einer Zelle auslesen ############################################################################
                for i in range(len(self.spalten_namen)):
                    zelle = self.sheet[
                        self.spalten_namen[(i + self.erste_leere_zelle_der_zeile_eins) - 1] + str(count_rows)].value

                    if not zelle:
                        # print(sheet[letters[i]+str(count_rows)].coordinate)
                        self.sheet[self.spalten_namen[(i + self.erste_leere_zelle_der_zeile_eins) - 1] + str(
                            count_rows)] = step_result

                        # Schreibe den letzten Status des Test-Steps
                        self.sheet['C' + str(count_rows)] = step_result
                        # Wenn der TestCase bereits in der Liste ist dann wird die Variabel auf True gesetzt
                        self.test_case_war_existent = True
                        break
                break

            count_rows = count_rows + 1

        # Wenn der Test-Case und der Schritt nicht in der Ergebnis-liste zu finden sind werden sie neu angelegt
        if self.test_case_war_existent == False:

            # print("Test Case wird neu erstellt  ")
            time.sleep(1)
            x = SchreibeTestCaseUndSchirttNameInFreieZeile(
                workbook=self.workbook,
                sheet=self.sheet,
            )
            x.schreibe_test_case_und_schritt_in_neue_zeile(
                methoden_name=self.methoden_name,
                schritt_oder_gesamter_testfall=schritt_oder_gesamter_testfall
            )

            time.sleep(2)

            for row in self.sheet.iter_rows(values_only=True):

                # Nachdem der Test-Fall und der Schritt angelegt wurden, wird erneut
                # überprüft in welcher Zeile der Spalte A sich der neue Testfall-Name un der TestSchritt befinden.
                # In diese Zeile wird dann das neuset Ergebnis eingetragen
                if row[0] == self.methoden_name and row[1] == schritt_oder_gesamter_testfall:

                    # Wert einer Zelle auslesen ########################################################################
                    for i in range(len(self.spalten_namen)):
                        zelle = self.sheet[
                            self.spalten_namen[(i + self.erste_leere_zelle_der_zeile_eins) - 1] +
                            str(count_empty_rows)].value

                        if not zelle:
                            # print(sheet[letters[i]+str(count_rows)].coordinate)
                            self.sheet[self.spalten_namen[(i + self.erste_leere_zelle_der_zeile_eins) - 1] + str(
                                count_empty_rows)] = step_result

                            # Schreibe den letzten Status des Test-Steps
                            self.sheet['C' + str(count_empty_rows)] = step_result
                            self.test_case_war_existent = True
                            break
                    break

                count_empty_rows = count_empty_rows + 1

        self.workbook.save(self.pfad_zur_test_ausfuehrungs_tabelle)

    def get_ergebnis_array(self):

        return self.ergebnis_array


